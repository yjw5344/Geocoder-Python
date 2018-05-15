import geocoder
import openpyxl
import time

## 엘셀파일 열기
filename = "address.xlsx"
exelFile = openpyxl.load_workbook(filename)

## 시트 설정
sheet = exelFile.worksheets[0]


## 데이터 가져오기
rowCount = 1
for row in sheet.rows:	

	## geocoder 호출
	g = geocoder.google(row[0].value)

	## 응답 실패 처리
	if g.status == "OVER_QUERY_LIMIT":
		while True:
			print("Fail ===> Try Again")
			time.sleep(0.5)
			g = geocoder.google(row[0].value)
			if g.status == "OK":
				break

	## cell 설정 [ B1 ~ B* : 위도 / C1 ~ C* : 경도]
	lat_cell = sheet.cell(row = rowCount, column = 2)
	lng_cell = sheet.cell(row = rowCount, column = 3)

	## 위도, 경도 변환
	geo = g.latlng

	## 변환 확인
	print(geo)

	## 데이터 추가
	lat_cell.value = geo[0]
	lng_cell.value = geo[1]
	rowCount = rowCount + 1

## 데이터 저장
exelFile.save("address.xlsx")




